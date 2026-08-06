import os
import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def pad_key(key):
    # 使用全角空格（\u3000）将所有键名对齐到4个中文字符宽度
    return key + '\u3000' * (4 - len(key))

def extract_court(text):
    """从PDF文本中提取法院名称"""
    # 策略1: 匹配"此致"后面的法院
    court_match = re.search(r"此致[\s\n]*([^\n]{2,}法院)", text)
    if court_match:
        return court_match.group(1).strip()

    # 策略2: 匹配文中出现的法院名称
    court_match = re.search(r"([\u4e00-\u9fa5]{2,}人民法院)", text)
    if court_match:
        return court_match.group(1).strip()

    return "未找到"

def process_folders(base_dir):
    # 存储所有案件信息，用于生成Excel
    all_cases = []

    # 遍历当前目录下的所有项
    for item in os.listdir(base_dir):
        folder_path = os.path.join(base_dir, item)

        # 判断是否为文件夹
        if os.path.isdir(folder_path):
            case_number = item
            pdf_path = os.path.join(folder_path, "强制执行申请书.pdf")
            txt_path = os.path.join(folder_path, "强制执行申请书.txt")

            # 如果文件夹内没有目标PDF，则跳过
            if not os.path.exists(pdf_path):
                continue

            print(f"正在处理文件夹: {item}")

            try:
                # 提取 PDF 文本
                with pdfplumber.open(pdf_path) as pdf:
                    text = ""
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"

                # 1. 基础信息提取（消除文本中的换行和空格）
                info_text = text.replace("\n", "").replace(" ", "").replace("\u3000", "")

                # 【关键修复1】圈定"被执行人"区域，避免提取到申请执行人的电话和信息
                respondent_match = re.search(r"被执行人：(.*?)请求事项", info_text)
                # 如果匹配成功则只在截取的内容里找，否则兜底用全文
                respondent_text = respondent_match.group(1) if respondent_match else info_text

                name_match = re.search(r"^([^，。、]+)", respondent_text)
                gender_match = re.search(r"性别：([^，。、]+)", respondent_text)

                # 兼容"电话："或"联系电话："
                phone_match = re.search(r"(?:联系)?电话：(\d+)", respondent_text)
                id_match = re.search(r"身份证号：([0-9xX]+)", respondent_text)
                address_match = re.search(r"住(.*?)(?:，|。|身份证)", respondent_text)

                # 提取法院信息
                court = extract_court(text)

                name = name_match.group(1) if name_match else "未找到"
                gender = gender_match.group(1) if gender_match else "未找到"
                phone = phone_match.group(1) if phone_match else "未找到"
                id_card = id_match.group(1) if id_match else "未找到"
                address = address_match.group(1) if address_match else "未找到"

                # 收集当前案件信息（用于Excel）
                case_info = {
                    "姓名": name,
                    "案号": case_number,
                    "电话": phone,
                    "身份证号": id_card,
                    "住址": address,
                    "法院": court
                }
                all_cases.append(case_info)

                # 2. 提取"执行请求"明细部分
                request_match = re.search(r"(1、.*?)(?=事实与理由|申请执行人|此致|\Z)", text, re.DOTALL)

                if request_match:
                    raw_request = request_match.group(1).strip()
                    request_text = re.sub(r'\s*\n\s*(?!\d+、)', '', raw_request)
                else:
                    request_text = "未找到执行请求明细内容"

                # 3. 按照要求格式化输出到txt
                sep = "     ->     "
                output_lines = [
                    f"{pad_key('案号')}{sep}{case_number}",
                    f"{pad_key('姓名')}{sep}{name}",
                    f"{pad_key('性别')}{sep}{gender}",
                    f"{pad_key('电话')}{sep}{phone}",
                    f"{pad_key('身份证号')}{sep}{id_card}",
                    f"{pad_key('住址')}{sep}{address}",
                    f"{pad_key('法院')}{sep}{court}",
                    "", "", "",
                    "执行请求",
                    "",
                    request_text
                ]

                # 4. 写入 txt 文件
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(output_lines))

                print(f"成功生成: {txt_path}")

            except Exception as e:
                print(f"处理 {item} 时出错: {e}")

    # ========== 新增：生成Excel汇总表 ==========
    if all_cases:
        generate_excel(base_dir, all_cases)
    else:
        print("未找到任何案件信息，跳过Excel生成")

def generate_excel(base_dir, all_cases):
    """生成类似截图的Excel汇总表"""
    excel_path = os.path.join(base_dir, "案件信息汇总表.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "案件汇总"

    # 定义样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    address_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 定义列（与截图一致）
    columns = ["姓名", "案号", "电话", "身份证号", "住址", "法院"]

    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, case in enumerate(all_cases, 2):
        ws.cell(row=row_idx, column=1, value=case["姓名"])
        ws.cell(row=row_idx, column=2, value=case["案号"])
        ws.cell(row=row_idx, column=3, value=case["电话"])
        ws.cell(row=row_idx, column=4, value=case["身份证号"])
        ws.cell(row=row_idx, column=5, value=case["住址"])
        ws.cell(row=row_idx, column=6, value=case["法院"])

        # 应用样式
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border
            # 住址列左对齐，其他居中
            if col_idx == 5:
                cell.alignment = address_alignment
            else:
                cell.alignment = cell_alignment

    # 调整列宽（与截图比例接近）
    ws.column_dimensions["A"].width = 10   # 姓名
    ws.column_dimensions["B"].width = 22   # 案号
    ws.column_dimensions["C"].width = 16   # 电话
    ws.column_dimensions["D"].width = 24   # 身份证号
    ws.column_dimensions["E"].width = 45  # 住址
    ws.column_dimensions["F"].width = 18   # 法院

    # 设置行高
    ws.row_dimensions[1].height = 28
    for row_idx in range(2, len(all_cases) + 2):
        ws.row_dimensions[row_idx].height = 22

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存
    wb.save(excel_path)
    print(f"\n✅ Excel汇总表已生成: {excel_path}")
    print(f"   共汇总 {len(all_cases)} 条案件信息")

if __name__ == "__main__":
    current_directory = os.getcwd()
    process_folders(current_directory)