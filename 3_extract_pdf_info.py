import os
import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def pad_key(key):
    # 使用全角空格（\u3000）将所有键名对齐到4个中文字符宽度
    return key + '\u3000' * (4 - len(key))

def extract_court(text):
    """从PDF文本中提取法院名称"""
    court_match = re.search(r"此致[\s\n]*([^\n]{2,}法院)", text)
    if court_match:
        return court_match.group(1).strip()

    court_match = re.search(r"([\u4e00-\u9fa5]{2,}人民法院)", text)
    if court_match:
        return court_match.group(1).strip()

    return "未找到"

def process_single_pdf(pdf_path, txt_path, case_number):
    """处理单个 PDF 文件，生成对应的 TXT 文件并返回案件信息字典"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

        info_text = text.replace("\n", "").replace(" ", "").replace("\u3000", "")

        # 兼容中文全角“：”和英文半角“:”
        respondent_match = re.search(r"被执行人[:：](.*?)请求事项", info_text)
        respondent_text = respondent_match.group(1) if respondent_match else info_text

        # 兼容中英文逗号
        name_match = re.search(r"^([^，,。、]+)", respondent_text)
        
        # PDF中可能没有“性别：”字样，直接匹配“男”或“女”
        gender_match = re.search(r"(?:性别[:：])?([男女])", respondent_text)
        
        # 兼容“联系电话:”、“电话：”
        phone_match = re.search(r"(?:联系)?电话[:：](\d+)", respondent_text)
        
        # 兼容“身份证号”和“身份证号码”，以及中英文冒号
        id_match = re.search(r"身份证号(?:码)?[:：]([0-9xX]+)", respondent_text)
        
        # 兼容“住”、“住址”、“地址”，截取到下一个逗号或句号
        address_match = re.search(r"(?:住址|地址|住)[:：]?([^，,。]+)", respondent_text)

        court = extract_court(text)
        name = name_match.group(1) if name_match else "未找到"
        gender = gender_match.group(1) if gender_match else "未找到"
        phone = phone_match.group(1) if phone_match else "未找到"
        id_card = id_match.group(1) if id_match else "未找到"
        address = address_match.group(1) if address_match else "未找到"

        case_info = {
            "姓名": name,
            "案号": case_number,
            "电话": phone,
            "身份证号": id_card,
            "住址": address,
            "法院": court
        }

        request_match = re.search(r"(1、.*?)(?=事实与理由|申请执行人|此致|\Z)", text, re.DOTALL)
        if request_match:
            raw_request = request_match.group(1).strip()
            request_text = re.sub(r'\s*\n\s*(?!\d+、)', '', raw_request)
        else:
            request_text = "未找到执行请求明细内容"

        sep = "     ->     "
        output_lines = [
            f"{pad_key('案号')}{sep}{case_number}",
            f"{pad_key('姓名')}{sep}{name}",
            f"{pad_key('性别')}{sep}{gender}",
            f"{pad_key('电话')}{sep}{phone}",
            f"{pad_key('身份证号')}{sep}{id_card}",
            f"{pad_key('住址')}{sep}{address}",
            f"{pad_key('法院')}{sep}{court}",
            "", "", "",
            "执行请求",
            "",
            request_text
        ]

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write("\n".join(output_lines))

        print(f"✅ 成功提取: {pdf_path}")
        return case_info

    except Exception as e:
        print(f"❌ 处理 {pdf_path} 时出错: {e}")
        return None

def process_folders(base_dir):
    subfolder_cases = []
    root_case = None

    print(f"当前锁定扫描路径: {base_dir}\n")

    # 1. 检查并处理当前目录(根目录)下的 PDF
    root_pdf_path = os.path.join(base_dir, "_强制执行申请书.pdf")
    root_txt_path = os.path.join(base_dir, "_强制执行申请书.txt")
    
    if os.path.exists(root_pdf_path) and os.path.isfile(root_pdf_path):
        print("发现当前目录(根目录)存在文件，正在处理...")
        case_number = os.path.basename(base_dir) or "当前目录"
        root_case = process_single_pdf(root_pdf_path, root_txt_path, case_number)

    # 2. 遍历当前目录下的所有子文件夹
    for item in os.listdir(base_dir):
        folder_path = os.path.join(base_dir, item)

        # 判断是否为文件夹
        if os.path.isdir(folder_path):
            case_number = item
            pdf_path = os.path.join(folder_path, "_强制执行申请书.pdf")
            txt_path = os.path.join(folder_path, "_强制执行申请书.txt")

            # 如果文件夹内没有目标PDF，则跳过
            if not os.path.exists(pdf_path) or not os.path.isfile(pdf_path):
                continue

            print(f"发现子文件夹 [{item}] 存在文件，正在处理...")
            case_info = process_single_pdf(pdf_path, txt_path, case_number)
            if case_info:
                subfolder_cases.append(case_info)

    # 3. 生成Excel汇总表逻辑控制
    if subfolder_cases:
        # 只要存在子目录的任务，就生成Excel。并将当前目录的数据汇总进去
        all_cases_for_excel = []
        if root_case:
            all_cases_for_excel.append(root_case)
        all_cases_for_excel.extend(subfolder_cases)
        
        generate_excel(base_dir, all_cases_for_excel)
    elif root_case:
        print("\n✅ 仅处理了当前目录的文件，未发现子目录任务，已按要求跳过生成 Excel 汇总表。")
    else:
        print("\n⚠️ 未在当前目录及子目录中找到任何 '_强制执行申请书.pdf' 文件。")

def generate_excel(base_dir, all_cases):
    """生成类似截图的Excel汇总表"""
    excel_path = os.path.join(base_dir, "案件信息汇总表.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "案件汇总"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    address_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    columns = ["姓名", "案号", "电话", "身份证号", "住址", "法院"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, case in enumerate(all_cases, 2):
        ws.cell(row=row_idx, column=1, value=case["姓名"])
        ws.cell(row=row_idx, column=2, value=case["案号"])
        ws.cell(row=row_idx, column=3, value=case["电话"])
        ws.cell(row=row_idx, column=4, value=case["身份证号"])
        ws.cell(row=row_idx, column=5, value=case["住址"])
        ws.cell(row=row_idx, column=6, value=case["法院"])

        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx == 5:
                cell.alignment = address_alignment
            else:
                cell.alignment = cell_alignment

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 18

    ws.row_dimensions[1].height = 28
    for row_idx in range(2, len(all_cases) + 2):
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A2"

    try:
        wb.save(excel_path)
        print(f"\n✅ Excel汇总表已生成: {excel_path}")
        print(f"   共汇总 {len(all_cases)} 条案件信息")
    except PermissionError:
        print(f"\n❌ [保存失败] 案件信息汇总表.xlsx 正在被其他软件（如Excel）占用！")
        print(f"   请先关闭该 Excel 文件，然后重新运行脚本。")

if __name__ == "__main__":
    # 【修复核心】使用 __file__ 锁定脚本绝对路径，彻底解决右键运行路径乱飘的问题
    current_directory = os.path.dirname(os.path.abspath(__file__))
    
    print("="*40)
    print(" PDF提取自动化脚本启动")
    print("="*40)
    
    process_folders(current_directory)
    
    print("\n" + "="*40)
    input(" 所有任务已执行完毕。按回车键(Enter)退出...")
