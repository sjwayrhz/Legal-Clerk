import os
import io
import re
import urllib.request
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader

# ========================================================================
# 结案申请书.pdf 生成模块（基于 结案申请书 模板样式复刻）
# 本版本不再依赖本地 assets 文件夹：
#   - 字体使用系统自带的"宋体"（simsun.ttc）
#   - 公章图片从公网直链实时下载
# ========================================================================

FONT_NAME = "SimSun"
SEAL_IMAGE_URL = "https://i.ibb.co/QFqp7nCx/seal-transparent.png"

# 常见系统中"宋体"字体文件的路径（按顺序尝试，找到第一个存在的即使用）
SIMSUN_CANDIDATE_PATHS = [
    r"C:\Windows\Fonts\simsun.ttc",
    r"C:\Windows\Fonts\SIMSUN.TTC",
    r"C:\Windows\Fonts\simsun.ttf",
]

_FONT_REGISTERED = False


def _ensure_font_registered():
    """确保宋体字体只注册一次（使用系统自带的 simsun.ttc，不再打包字体文件）"""
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return

    font_path = next((p for p in SIMSUN_CANDIDATE_PATHS if os.path.exists(p)), None)
    if not font_path:
        raise FileNotFoundError(
            "未能在本机找到系统自带的宋体字体文件（默认路径: C:\\Windows\\Fonts\\simsun.ttc）。\n"
            "请确认本机为 Windows 系统且已安装宋体，或修改脚本中 SIMSUN_CANDIDATE_PATHS 指向正确的字体路径。"
        )

    # simsun.ttc 是字体合集文件，index 0 对应"宋体"（index 1 是"新宋体"）
    subfont_index = 0 if font_path.lower().endswith(".ttc") else 0
    pdfmetrics.registerFont(TTFont(FONT_NAME, font_path, subfontIndex=subfont_index))
    _FONT_REGISTERED = True


_SEAL_IMAGE_BYTES = None
_SEAL_DOWNLOAD_ATTEMPTED = False


def _get_seal_image_bytes():
    """从公网直链下载公章图片（整个脚本运行期间只下载一次，缓存在内存中）"""
    global _SEAL_IMAGE_BYTES, _SEAL_DOWNLOAD_ATTEMPTED
    if _SEAL_DOWNLOAD_ATTEMPTED:
        return _SEAL_IMAGE_BYTES

    _SEAL_DOWNLOAD_ATTEMPTED = True
    try:
        req = urllib.request.Request(
            SEAL_IMAGE_URL, headers={"User-Agent": "Mozilla/5.0"}
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            _SEAL_IMAGE_BYTES = resp.read()
    except Exception as e:
        print(f"⚠️ 公章图片下载失败（{e}），生成的结案申请书.pdf 中将不含公章图案。")
        _SEAL_IMAGE_BYTES = None

    return _SEAL_IMAGE_BYTES


# ---- 版式参数（根据 结案申请书 模板量取，单位:pt，页面尺寸约等于A4）----
PAGE_W, PAGE_H = 595.5, 842.0
BODY_SIZE = 14.1
TITLE_SIZE = 22
LINE_HEIGHT = 24        # 同一段落内的行距
PARA_GAP = 28.2         # 段落之间的额外间距（段前距）
X_FIRST = 115.3         # 段落首行缩进后的起始X坐标
X_CONT = 86.3           # 段落非首行（续行）的起始X坐标
RIGHT_EDGE = 540.5      # 正文右边界
TOP_START_Y_FITZ = 103.8  # 第一段首行相对于页面顶部的偏移（PDF常见坐标系，从上往下量取）

AVOID_LINE_START = set("，。、；：！？）】」』”’,.;:!?)")


def _fitz_top_to_pdf_baseline(fitz_y0, size=BODY_SIZE):
    """将“从页面顶部往下量取的Y坐标”换算为reportlab所需的基线Y坐标（从页面底部往上量取）"""
    return PAGE_H - fitz_y0 - size * 0.83


def _wrap_paragraph(text, font_name, font_size, x_first, x_cont, right_edge):
    """按可用宽度对整段中文文本进行折行（逐字符计算宽度），并做简单的标点悬挂处理"""
    first_width = right_edge - x_first
    cont_width = right_edge - x_cont
    chars = list(text)
    lines = []
    idx = 0
    first_line = True
    n = len(chars)
    while idx < n:
        width_limit = first_width if first_line else cont_width
        cur = ""
        cur_w = 0.0
        while idx < n:
            ch = chars[idx]
            w = pdfmetrics.stringWidth(ch, font_name, font_size)
            if cur and cur_w + w > width_limit:
                break
            cur += ch
            cur_w += w
            idx += 1
        # 避免标点符号被挤到下一行行首（悬挂标点）
        while idx < n and chars[idx] in AVOID_LINE_START:
            cur += chars[idx]
            idx += 1
        lines.append(cur)
        first_line = False
    return lines


def _draw_justified_line(c, chars, x, y, font_name, font_size, target_width):
    """两端对齐绘制一行文字（通过在字符间插入等量空隙撑满整行宽度）"""
    if len(chars) <= 1:
        c.setFont(font_name, font_size)
        c.drawString(x, y, chars)
        return
    total_w = sum(pdfmetrics.stringWidth(ch, font_name, font_size) for ch in chars)
    gap_count = len(chars) - 1
    extra = max(0.0, target_width - total_w)
    extra_per_gap = extra / gap_count if gap_count else 0.0
    c.setFont(font_name, font_size)
    cx = x
    for i, ch in enumerate(chars):
        c.drawString(cx, y, ch)
        cx += pdfmetrics.stringWidth(ch, font_name, font_size) + extra_per_gap


def _draw_paragraph(c, text, y_baseline, font_name=FONT_NAME, font_size=BODY_SIZE,
                     x_first=X_FIRST, x_cont=X_CONT, right_edge=RIGHT_EDGE,
                     line_height=LINE_HEIGHT, justify=True):
    """绘制一个段落（自动折行，除最后一行外两端对齐），返回最后一行的基线Y坐标"""
    lines = _wrap_paragraph(text, font_name, font_size, x_first, x_cont, right_edge)
    first_width = right_edge - x_first
    cont_width = right_edge - x_cont
    y = y_baseline
    for i, line in enumerate(lines):
        x = x_first if i == 0 else x_cont
        target_width = first_width if i == 0 else cont_width
        is_last = (i == len(lines) - 1)
        if justify and not is_last:
            _draw_justified_line(c, line, x, y, font_name, font_size, target_width)
        else:
            c.setFont(font_name, font_size)
            c.drawString(x, y, line)
        if i < len(lines) - 1:
            y -= line_height
    return y


def generate_closing_application_pdf(output_path, respondent_text, name, case_number,
                                      applicant_court="滨州市中级人民法院"):
    """
    生成《结案申请书》PDF（以用户提供的模板样式复刻）。

    :param output_path: 输出的 结案申请书.pdf 路径
    :param respondent_text: 从《强制执行申请书》中提取的完整被执行人信息原文
                             （形如："袁野，性别：男，民族：汉族，1996年10月27日出生，
                             住xxx，身份证号：xxx，联系电话：xxx。"）
    :param name: 被执行人姓名（用于替换模板正文中出现的"袁野"）
    :param case_number: 案号（用于替换模板中的"（2024）衢仲字第1-3816号"，即子文件夹名称）
    :param applicant_court: 此致的法院名称，默认与模板保持一致
    """
    _ensure_font_registered()

    c = canvas.Canvas(output_path, pagesize=(PAGE_W, PAGE_H))

    # 标题
    title = "结案申请书"
    c.setFont(FONT_NAME, TITLE_SIZE)
    title_w = pdfmetrics.stringWidth(title, FONT_NAME, TITLE_SIZE)
    title_x = (PAGE_W - title_w) / 2
    title_y = _fitz_top_to_pdf_baseline(67.5, TITLE_SIZE)
    c.drawString(title_x, title_y, title)

    y = _fitz_top_to_pdf_baseline(TOP_START_Y_FITZ)

    paragraphs = [
        "申请执行人：河南中原消费金融股份有限公司，统一社会信用代码：91410000MA403TWC10，"
        "地址：河南省郑州市郑东新区明理路与白佛南路交叉口中原金融产业园8号楼。",
        "法定代表人：马景鹏",
        f"被执行人：{respondent_text}",
        "请求事项：",
        f"贵院受理的申请执行人河南中原消费金融股份有限公司与被执行人{name}金融借款合同纠纷一案，"
        f"申请结案，并解除对被执行人名下的强制措施。",
        f"执行依据：{case_number}",
        "事实与理由：",
        f"贵院受理的申请人河南中原消费金融股份有限公司与被执行人{name}金融借款合同纠纷一案，"
        f"执行依据为{case_number}仲裁裁决书，现被执行人已履行完毕，特向贵院申请结案，望贵院予以准许。",
        "此致",
        applicant_court,
    ]

    for p in paragraphs:
        last_line_y = _draw_paragraph(c, p, y)
        y = last_line_y - PARA_GAP

    # 落款（申请人 + 公章），位置与模板保持一致
    sig_text = "申请人：河南中原消费金融股份有限公司"
    sig_y = _fitz_top_to_pdf_baseline(674.2)
    c.setFont(FONT_NAME, BODY_SIZE)
    c.drawString(281.5, sig_y, sig_text)

    seal_bytes = _get_seal_image_bytes()
    if seal_bytes:
        seal_x, seal_y_top_fitz, seal_x2, seal_y_bottom_fitz = 370.68, 629.4, 503.04, 760.8
        seal_w = seal_x2 - seal_x
        seal_h = seal_y_bottom_fitz - seal_y_top_fitz
        seal_pdf_y = PAGE_H - seal_y_bottom_fitz
        c.drawImage(ImageReader(io.BytesIO(seal_bytes)), seal_x, seal_pdf_y,
                    width=seal_w, height=seal_h, mask='auto')

    # 日期行（模板中未填写具体日期，保持空白，与模板一致）
    date_y = _fitz_top_to_pdf_baseline(702.4)
    c.setFont(FONT_NAME, BODY_SIZE)
    c.drawString(453.1, date_y, "年")
    c.drawString(489.7, date_y, "月")
    c.drawString(526.3, date_y, "日")

    c.save()


# ========================================================================
# 原有信息提取逻辑
# ========================================================================

def pad_key(key):
    # 使用全角空格（\u3000）将所有键名对齐到4个中文字符宽度
    return key + '\u3000' * (4 - len(key))

def extract_court(text):
    """从PDF文本中提取法院名称"""
    court_match = re.search(r"此致[\s\n]*([^\n]{2,}法院)", text)
    if court_match:
        return court_match.group(1).strip()

    court_match = re.search(r"([\u4e00-\u9fa5]{2,}人民法院)", text)
    if court_match:
        return court_match.group(1).strip()

    return "未找到"

def process_single_pdf(pdf_path, txt_path, case_number):
    """处理单个 PDF 文件，生成对应的 TXT 文件、结案申请书.pdf 并返回案件信息字典"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

        info_text = text.replace("\n", "").replace(" ", "").replace("\u3000", "")

        respondent_match = re.search(r"被执行人：(.*?)请求事项", info_text)
        respondent_text = respondent_match.group(1) if respondent_match else info_text

        name_match = re.search(r"^([^，。、]+)", respondent_text)
        gender_match = re.search(r"性别：([^，。、]+)", respondent_text)
        phone_match = re.search(r"(?:联系)?电话：(\d+)", respondent_text)
        id_match = re.search(r"身份证号：([0-9xX]+)", respondent_text)
        address_match = re.search(r"住(.*?)(?:，|。|身份证)", respondent_text)

        court = extract_court(text)
        name = name_match.group(1) if name_match else "未找到"
        gender = gender_match.group(1) if gender_match else "未找到"
        phone = phone_match.group(1) if phone_match else "未找到"
        id_card = id_match.group(1) if id_match else "未找到"
        address = address_match.group(1) if address_match else "未找到"

        case_info = {
            "姓名": name,
            "案号": case_number,
            "电话": phone,
            "身份证号": id_card,
            "住址": address,
            "法院": court
        }

        request_match = re.search(r"(1、.*?)(?=事实与理由|申请执行人|此致|\Z)", text, re.DOTALL)
        if request_match:
            raw_request = request_match.group(1).strip()
            request_text = re.sub(r'\s*\n\s*(?!\d+、)', '', raw_request)
        else:
            request_text = "未找到执行请求明细内容"

        sep = "     ->     "
        output_lines = [
            f"{pad_key('案号')}{sep}{case_number}",
            f"{pad_key('姓名')}{sep}{name}",
            f"{pad_key('性别')}{sep}{gender}",
            f"{pad_key('电话')}{sep}{phone}",
            f"{pad_key('身份证号')}{sep}{id_card}",
            f"{pad_key('住址')}{sep}{address}",
            f"{pad_key('法院')}{sep}{court}",
            "", "", "",
            "执行请求",
            "",
            request_text
        ]

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write("\n".join(output_lines))

        print(f"✅ 成功提取: {pdf_path}")

        # ---- 生成 结案申请书.pdf ----
        if respondent_match and name_match:
            try:
                closing_pdf_path = os.path.join(os.path.dirname(pdf_path), "结案申请书.pdf")
                generate_closing_application_pdf(
                    output_path=closing_pdf_path,
                    respondent_text=respondent_text,
                    name=name,
                    case_number=case_number
                )
                print(f"   ✅ 已生成结案申请书: {closing_pdf_path}")
            except Exception as gen_err:
                print(f"   ❌ 生成结案申请书.pdf 时出错: {gen_err}")
        else:
            print(f"   ⚠️ 未能可靠提取被执行人信息，已跳过生成结案申请书.pdf: {pdf_path}")

        return case_info

    except Exception as e:
        print(f"❌ 处理 {pdf_path} 时出错: {e}")
        return None

def process_folders(base_dir):
    subfolder_cases = []
    root_case = None

    print(f"当前锁定扫描路径: {base_dir}\n")

    # 1. 检查并处理当前目录(根目录)下的 PDF
    root_pdf_path = os.path.join(base_dir, "强制执行申请书.pdf")
    root_txt_path = os.path.join(base_dir, "强制执行申请书.txt")
    
    if os.path.exists(root_pdf_path) and os.path.isfile(root_pdf_path):
        print("发现当前目录(根目录)存在文件，正在处理...")
        case_number = os.path.basename(base_dir) or "当前目录"
        root_case = process_single_pdf(root_pdf_path, root_txt_path, case_number)

    # 2. 遍历当前目录下的所有子文件夹
    for item in os.listdir(base_dir):
        folder_path = os.path.join(base_dir, item)

        # 判断是否为文件夹
        if os.path.isdir(folder_path):
            case_number = item
            pdf_path = os.path.join(folder_path, "强制执行申请书.pdf")
            txt_path = os.path.join(folder_path, "强制执行申请书.txt")

            # 如果文件夹内没有目标PDF，则跳过
            if not os.path.exists(pdf_path) or not os.path.isfile(pdf_path):
                continue

            print(f"发现子文件夹 [{item}] 存在文件，正在处理...")
            case_info = process_single_pdf(pdf_path, txt_path, case_number)
            if case_info:
                subfolder_cases.append(case_info)

    # 3. 生成Excel汇总表逻辑控制
    if subfolder_cases:
        # 只要存在子目录的任务，就生成Excel。并将当前目录的数据汇总进去
        all_cases_for_excel = []
        if root_case:
            all_cases_for_excel.append(root_case)
        all_cases_for_excel.extend(subfolder_cases)
        
        generate_excel(base_dir, all_cases_for_excel)
    elif root_case:
        print("\n✅ 仅处理了当前目录的文件，未发现子目录任务，已按要求跳过生成 Excel 汇总表。")
    else:
        print("\n⚠️ 未在当前目录及子目录中找到任何 '强制执行申请书.pdf' 文件。")

def generate_excel(base_dir, all_cases):
    """生成类似截图的Excel汇总表"""
    excel_path = os.path.join(base_dir, "案件信息汇总表.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "案件汇总"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    address_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    columns = ["姓名", "案号", "电话", "身份证号", "住址", "法院"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, case in enumerate(all_cases, 2):
        ws.cell(row=row_idx, column=1, value=case["姓名"])
        ws.cell(row=row_idx, column=2, value=case["案号"])
        ws.cell(row=row_idx, column=3, value=case["电话"])
        ws.cell(row=row_idx, column=4, value=case["身份证号"])
        ws.cell(row=row_idx, column=5, value=case["住址"])
        ws.cell(row=row_idx, column=6, value=case["法院"])

        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx == 5:
                cell.alignment = address_alignment
            else:
                cell.alignment = cell_alignment

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 18

    ws.row_dimensions[1].height = 28
    for row_idx in range(2, len(all_cases) + 2):
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A2"

    try:
        wb.save(excel_path)
        print(f"\n✅ Excel汇总表已生成: {excel_path}")
        print(f"   共汇总 {len(all_cases)} 条案件信息")
    except PermissionError:
        print(f"\n❌ [保存失败] 案件信息汇总表.xlsx 正在被其他软件（如Excel）占用！")
        print(f"   请先关闭该 Excel 文件，然后重新运行脚本。")

if __name__ == "__main__":
    # 【修复核心】使用 __file__ 锁定脚本绝对路径，彻底解决右键运行路径乱飘的问题
    current_directory = os.path.dirname(os.path.abspath(__file__))
    
    print("="*40)
    print(" PDF提取自动化脚本启动")
    print("="*40)
    
    process_folders(current_directory)
    
    print("\n" + "="*40)
    input(" 所有任务已执行完毕。按回车键(Enter)退出...")
